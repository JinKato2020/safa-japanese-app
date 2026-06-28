# -*- coding: utf-8 -*-
import openpyxl, json, re, os
SRC=r"C:/Users/jwpsa/Documents/desktop/claude/safa-japanese-app/読解.xlsx"
OUT=r"C:/Users/jwpsa/Documents/desktop/claude/safa-japanese-app/src/data/reading.json"
wb=openpyxl.load_workbook(SRC, read_only=True, data_only=True)

CAT_LABEL={"T":"内容理解(短文)","C":"内容理解(中文)","L":"内容理解(長文)","J":"情報検索"}
CAT_ORDER=["T","C","L","J"]
LEVELS=["N5","N4","N3"]

# data[level][cat] = list of scripts; each script = {scriptId,passage,charCount,items:[...]}
data={lv:{c:[] for c in CAT_ORDER} for lv in LEVELS}
script_index={}  # (level,cat,scriptId) -> script dict

def is_placeholder(t):
    if not t: return True
    s=str(t).strip()
    return s.startswith("（同じ本文") or s.startswith("(同じ本文")

for name in wb.sheetnames:
    if name=="説明": continue
    ws=wb[name]
    rows=list(ws.iter_rows(values_only=True))
    for r in rows[2:]:
        rid=r[0]
        if not rid or not str(rid).startswith("N"): continue
        rid=str(rid).strip()
        parts=rid.split("-")           # N5-D-T-001
        level=parts[0]; cat=parts[2]
        scriptId=str(r[1]).strip() if r[1] else rid
        passage=r[3]; charcnt=r[4]
        question=str(r[5]).strip() if r[5] else ""
        choices=[str(r[i]).strip() for i in (6,7,8,9) if r[i] is not None and str(r[i]).strip()!=""]
        answer=str(r[6]).strip() if r[6] else (choices[0] if choices else "")  # 選択肢1=正解
        difficulty=str(r[11]).strip() if r[11] else ""
        aim=str(r[12]).strip() if r[12] else ""
        figure=str(r[14]).strip() if r[14] else ""
        note=str(r[15]).strip() if r[15] else ""
        key=(level,cat,scriptId)
        if key not in script_index:
            sd={"scriptId":scriptId,
                "passage": "" if is_placeholder(passage) else str(passage).strip(),
                "charCount": charcnt if isinstance(charcnt,int) else None,
                "items":[]}
            script_index[key]=sd
            data[level][cat].append(sd)
        sd=script_index[key]
        # fill passage if earlier was placeholder
        if not sd["passage"] and not is_placeholder(passage):
            sd["passage"]=str(passage).strip()
        sd["items"].append({
            "id":rid,"question":question,"choices":choices,"answer":answer,
            "difficulty":difficulty,"aim":aim,
            **({"figure":figure} if figure and figure!="不要" else {}),
            **({"note":note} if note else {}),
        })

out={
  "generatedFrom":"読解.xlsx",
  "levels":LEVELS,
  "categoryOrder":CAT_ORDER,
  "categoryLabels":CAT_LABEL,
  "data":data,
}
# stats + sanity
stats={}
for lv in LEVELS:
    for c in CAT_ORDER:
        scripts=data[lv][c]
        nq=sum(len(s["items"]) for s in scripts)
        nopass=sum(1 for s in scripts if not s["passage"])
        if scripts:
            stats[f"{lv}/{c}"]=f"{len(scripts)}本/{nq}問" + (f" ⚠passage欠{nopass}" if nopass else "")
os.makedirs(os.path.dirname(OUT),exist_ok=True)
json.dump(out, open(OUT,"w",encoding="utf-8"), ensure_ascii=False, indent=1)
print("written",OUT, os.path.getsize(OUT),"bytes")
for k,v in stats.items(): print(" ",k,v)
