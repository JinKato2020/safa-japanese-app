# -*- coding: utf-8 -*-
# 短文.xlsx / 長文.xlsx（各カテゴリー＝シート）→ src/data/content.json（ツリー）
# IDは各xlsxの ID 列をそのまま使う（採番済み: S-<code>-NNN / L-<code>-NNN）。
# ツリー: タブ ＞ カテゴリー ＞ (サブテーマ) ＞ items。カイの物語は カテゴリー=カイの物語・サブテーマ=区分 でまとまる。
import openpyxl, json, os
APP = r"C:/Users/jwpsa/Documents/desktop/claude/safa-japanese-app"
OUT = os.path.join(APP, "src", "data", "content.json")
FILES = [("短文", "短文.xlsx"), ("長文", "長文.xlsx")]
ORDER = {
 "短文": ["生活力","文化","本音講座","日本と世界の違い","使える一言","グルメ","ことわざ","カタカナ語","擬音語"],
 "長文": ["カイの物語","ラジオトーク","ミステリー","ドキュメンタリー","昔話"],
}
TABS = ["短文","長文"]
data = {t: [] for t in TABS}

def child(lst, name):
    for n in lst:
        if n["name"] == name: return n
    n = {"name": name, "children": []}; lst.append(n); return n

for tab, fn in FILES:
    wb = openpyxl.load_workbook(os.path.join(APP, fn), read_only=True, data_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]; rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        ci = {h: i for i, h in enumerate(rows[0])}
        for r in rows[1:]:
            if not r[ci["ID"]]: continue
            cat = str(r[ci["カテゴリー"]]).strip()
            sv = r[ci["サブテーマ"]]
            sub = "" if sv is None else str(sv).strip()
            node = child(data[tab], cat)
            if sub and sub != "None":
                node = child(node["children"], sub)
            node.setdefault("items", [])
            node["items"].append({
                "id": str(r[ci["ID"]]).strip(),
                "title": str(r[ci["タイトル"]]).strip(),
                "form": str(r[ci["形式"]]).strip(),
                "level": str(r[ci["対象レベル(非表示)"]]).strip() if r[ci["対象レベル(非表示)"]] else "",
                "text": str(r[ci["本文"]]).strip(),
                "en": str(r[ci["英訳"]]).strip() if r[ci["英訳"]] else "",
                "key": str(r[ci["キー表現"]]).strip() if r[ci["キー表現"]] else "",
                "point": str(r[ci["ポイント"]]).strip() if r[ci["ポイント"]] else "",
                "note": str(r[ci["備考(音声)"]]).strip() if r[ci["備考(音声)"]] else "",
            })

# children 空ノードを削除
def prune(n):
    for c in n.get("children", []): prune(c)
    if "children" in n and not n["children"]: del n["children"]
for t in TABS:
    for c in data[t]: prune(c)
# カテゴリー順に並べ替え
for t in TABS:
    data[t].sort(key=lambda n: ORDER[t].index(n["name"]) if n["name"] in ORDER[t] else 999)

json.dump({"tabs": TABS, "data": data}, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("written", OUT)
def walk(n, d=0):
    has = "items" in n
    print("  "*d + "・" + n["name"] + (f" [{len(n['items'])}本]" if has else ""))
    if has:
        for it in n["items"][:1]: print("  "*(d+1) + "例ID:", it["id"], it["title"])
    for c in n.get("children", []): walk(c, d+1)
for t in TABS:
    print("["+t+"]")
    for c in data[t]: walk(c, 1)
